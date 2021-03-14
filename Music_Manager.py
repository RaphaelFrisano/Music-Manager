import time
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import spotipy
from spotipy import util
from spotipy.oauth2 import SpotifyClientCredentials
from spotipy.oauth2 import SpotifyOAuth

def get_playlist_tracks(sp, username, playlist_id):
    results = sp.user_playlist_tracks(username,playlist_id)
    tracks = results['items']
    while results['next']:
        results = sp.next(results)
        tracks.extend(results['items'])
    return tracks

#Set up Spotify Connection
client_id = '4b3eebc429cd462c9820a2e20fd89ef5'
f = open("secret.txt", "r")
client_secret = str(f.read())
username = '16r49f73ryoeuabwxqwgpimzs'
scope = 'user-library-read user-library-modify playlist-modify-public playlist-modify-private'
redirect_uri='http://localhost:8888/callback'
client_credentials_manager = SpotifyClientCredentials(client_id=client_id, client_secret=client_secret)
#sp = spotipy.Spotify(client_credentials_manager=client_credentials_manager)
auth_manager = SpotifyOAuth(client_id=client_id, client_secret=client_secret, scope=scope, redirect_uri=redirect_uri)
sp = spotipy.Spotify(auth_manager=auth_manager)
token = util.prompt_for_user_token(username, scope, client_id, client_secret, redirect_uri)
if token:
    sp = spotipy.Spotify(auth=token)
else:
    print("Can't get token for", username)


# Initialisation
favouritesTracks = get_playlist_tracks(sp, username, "1W62xrI7rdqkrkT0xgbwTM")
DanceVibesPlaylist = "47wDz8QxSAsDOLB8c8HdCd"
HopPopVibesPlaylist = "tT9IzMkySrGHwNmhBBnB5w"
ClassicVibesPlaylist = "31159C1hoG2ZqZxtTLvBk2"
EasyVibesPlaylist = "7Gq1sXY7RZ3LSKm86bPn7v"
RockyVibesPlaylist = "67sSpThGjgTPiRnZ1S8GIW"
VibeCheckPlaylist = "4GtrzqGPZdCKA29WQWlRdJ"
DanceVibes = []
HopPopVibes = []
ClassicVibes = []
EasyVibes = []
RockyVibes = []
VibeCheck = []
Songnum = 0
SongnumEnd = (len(favouritesTracks) - 1)


# Manage Music into Playlists
while Songnum < SongnumEnd:
    Songnum = Songnum + 1
    ParentGenres = "Dance/Electronic"
    genres = ParentGenres.split(",")
    for genre in genres:
        if genre in "Dance/Electronic":
            DanceVibes.append(favouritesTracks[Songnum]["track"]["id"])

        elif genre in "Pop" or genre in "HipHop":
            HopPopVibes.append(favouritesTracks[Songnum]["track"]["id"])

        elif genre in "Classical":
            ClassicVibes.append(favouritesTracks[Songnum]["track"]["id"])

        elif genre in "Easy Listening":
            EasyVibes.append(favouritesTracks[Songnum]["track"]["id"])

        elif genre in "Rock" or genre in "Metal":
            RockyVibes.append(favouritesTracks[Songnum]["track"]["id"])
        
        else:
            VibeCheck.append(favouritesTracks[Songnum]["track"]["id"])
    
sp.playlist_add_items("47wDz8QxSAsDOLB8c8HdCd", DanceVibes)
sp.playlist_add_items("tT9IzMkySrGHwNmhBBnB5w", HopPopVibes)
sp.playlist_add_items("31159C1hoG2ZqZxtTLvBk2", ClassicVibes)
sp.playlist_add_items("7Gq1sXY7RZ3LSKm86bPn7v", EasyVibes)
sp.playlist_add_items("67sSpThGjgTPiRnZ1S8GIW", RockyVibes)
sp.playlist_add_items("4GtrzqGPZdCKA29WQWlRdJ", VibeCheck)


'''
#Connect to playlist
trackslist = sp.current_user_saved_tracks(limit=50, offset=0)

##Open Excel
#xlxwb = load_workbook(os.getcwd() + "/SongsList.xlsx")
#xlx = xlxwb['Tabelle1']
#
## Get songs
#excelStartFile = open(os.getcwd() + '/startline.txt', 'r+')
#row = excelStartFile.read()
#row = [int(s) for s in row.split() if s.isdigit()] # Get only numbers out of File
#row = row[0] # Get first number

trackslist = sp.current_user_saved_tracks(limit=50, offset=0)
offsetvar = 0
while(offsetvar <= trackslist['total']):
    
    trackslist = sp.current_user_saved_tracks(limit=50, offset=offsetvar)
    for songdict in trackslist['items']:
        # Get song Info
        songtitle = songdict['track']['name']
        artist = songdict['track']['artists'][0]['name']

        print("=======================|>")
        print("Title - " + songtitle)
        print("Artist - " + artist)

    
    offsetvar = offsetvar + 50

    # Add song to Excel
    xlx.cell(row = row, column = 1).value = str(artist)
    xlx.cell(row = row, column = 2).value = str(songtitle)
    xlx.cell(row = row, column = 3).value = str(time.asctime())

    row = row + 1
'''